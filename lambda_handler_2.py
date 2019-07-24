import boto3, logging, psycopg2
from openpyxl import load_workbook
from configparser import ConfigParser
from io import BytesIO

logger = logging.getLogger()
logger.setLevel(logging.INFO)
s3_cli = boto3.client('s3')

DB_CONFIG = '/tmp/config.ini'

def begin(event,context):
    logger.info(event)
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']
    getDBConfig(bucket)
    insert_data = openExcel(bucket,key)
    insertRDS(insert_data)

def getDBConfig(bucket):
    responseObj = s3_cli.get_object(Bucket=bucket, Key=DB_CONFIG.rsplit('/')[-1])
    cfg = responseObj['Body'].read()
    with open(DB_CONFIG,'wb') as config:
        config.write(cfg)

def openExcel(bucket,key):
    responseObj = s3_cli.get_object(Bucket=bucket, Key=key)
    excelBinaryData = responseObj['Body'].read()
    '''use BytesIO to load binary data, instead of writing to disk'''
    wb = load_workbook(BytesIO(excelBinaryData))
    print('Opened Excel Spreadsheet')
    sheet = wb.sheetnames[0]
    active_sheet = wb[sheet]
    main = []
    '''create a list of  tuple lists by extracting all rows and columns from spreadsheet'''
    for row in active_sheet.values:
        save_list = []
        for column in row:
            record = column
            save_list.append(tuple([record]))
        main.append(save_list)
    print(main)
    return main

def config(filename=DB_CONFIG, section='rds'):
    '''parse conf file '''
    parser = ConfigParser()
    '''read the config'''
    parser.read(filename)
    '''get the section wanted'''
    db = {}
    if parser.has_section(section):
        print('Reading db config')
        params = parser.items(section)
        for param in params:
            db[param[0]] = param[1]   #key value structure from file
    else:
        raise Exception('Section {0} is not found in {1} file'.format(section,filename))
    return db

def insertRDS(data_list):
    # %s denotes string value
    sql = "Insert INTO tx(City,State,County,Established,Alias,Initials,Email1,Email2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
    connection = None
    try:
        params = config()
        logger.info(params)
        connection = psycopg2.connect(**params)
        cursor = connection.cursor()
        for i in range(len(data_list)):
            cursor.execute(sql, data_list[i])
        connection.commit()
        cursor.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if connection is not None:
            connection.close()


#AWS Lambda practice: Triggering Lambda functions - part 2
#Function is enhanced to read multi column/row spreadsheet
#Triggered by the upload of a spreadsheet file (.xlsx) to s3
#The data is parsed and then written to (POSTGRE) RDS table in AWS
#Elliott Arnold 7-23-19
#si3mshady

#https://en.wikipedia.org/wiki/List_of_counties_in_Texas